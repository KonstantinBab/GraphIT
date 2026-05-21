# parse_class_p.py — Парсинг PDF через Ollama API (requests, без библиотеки ollama)
#
# Работает с любой vision-моделью: devstral, qwen3-vl, и т.д.
# Прямые HTTP запросы к Ollama REST API — полный контроль.

import re
import json
import base64
import requests
import pandas as pd
import numpy as np
from io import BytesIO, StringIO
from typing import List, Optional
import cv2
from PIL import Image
from pdf2image import convert_from_path
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock


class ConstructionBillOCR:
    OCR_PROMPT = """
    Ты OCR. Перепиши таблицу с изображения в markdown.

    Правила:
    1. НЕ объединяй строки. НЕ изменяй текст. НЕ пропускай строки. НЕ добавляй от себя.
    2. Каждая видимая строка на изображении = отдельная строка в таблице.
    3. Штампы и подписи внизу страницы не переписывай.

    ═══════════════════════════════════════════════════════════════
    ТИП 1 (Сметы, ГЭСН) — определяем по наличию цифр в заголовках
    ═══════════════════════════════════════════════════════════════

    Заголовок раздела:
    - ПУСТОЙ № п/п + пустые Ед. изм./Кол-во
    - Текст начинается с цифры: "1 Металлоконструкции", "2. Бурение"
    - Оберни в **двойные звёздочки**
    - Пример: | | **1 Металлоконструкции** | | |

    Подзаголовок:
    - ЗАПОЛНЕН № п/п (цифра) + пустые Ед. изм./Кол-во
    - Текст БЕЗ цифры в начале: "Антикоррозионное покрытие"
    - Оберни в *одинарные звёздочки*
    - Пример: | 1.1 | *Антикоррозионное покрытие* | | |

    ═══════════════════════════════════════════════════════════════
    ТИП 2 (Газопроводы, линейные объекты) — определяем по визуальному форматированию
    ═══════════════════════════════════════════════════════════════

    ВАЖНО — ВИЗУАЛЬНОЕ ФОРМАТИРОВАНИЕ:
    Обращай внимание на оформление текста на изображении:
    - Если текст ВЫДЕЛЕН ЖИРНЫМ шрифтом И ПОДЧЁРКНУТ → это раздел (заголовок или подраздел)

    Заголовок раздела:
    - Визуально: жирный + подчёркнутый
    - ПУСТОЙ № п/п
    - Оберни такой текст в **двойные звёздочки**
    - Пример: | | **Прокладка газопровода DN1400** | | |

    Подраздел:
    - Визуально: жирный + подчёркнутый
    - ЕСТЬ № п/п
    - Оберни такой текст в *одинарные звёздочки*
    - Пример: | 1 | *Ведомость объемов работ* | | |

    ═══════════════════════════════════════════════════════════════
    ОБЩИЕ ПРАВИЛА ДЛЯ ВСЕХ ТИПОВ
    ═══════════════════════════════════════════════════════════════

    Что НЕ выделять (обычные позиции):
    - Строки с заполненными Ед. изм. или Кол-во
    - Подпункты с точкой: "1.1 Монтаж" (если есть Ед. изм./Кол-во)
    - Размеры и единицы: "100мм труба", "50кг мешок"
    - Технические марки: "В25 бетон", "А500С арматура"

    Примеры вывода:
    | | **1 Металлоконструкции** | | |
    | 1.1 | *Антикоррозионное покрытие* | | |
    | 1.1.1 | Монтаж колонн | шт | 10 |
    | | **Прокладка газопровода DN1400** | | |
    | 1 | *Ведомость объемов работ* | | |
    | 1.1 | Сварка стыков | стык | 50 |

    Формат таблицы:
    | № п/п | Наименование | Ед. изм. | Кол-во |
    |-------|--------------|----------|--------|
    """.strip()

    def __init__(
        self,
        pdf_path: str,
        output_xlsx: Optional[str] = None,
        poppler_path: Optional[str] = None,
        ollama_url: str = "http://localhost:11434",
        ollama_model: str = "devstral-small-2:24b",
        dpi: int = 400,
        preprocess_images: bool = True,
        ocr_prompt: Optional[str] = None,
        progress_callback=None,
        max_parallel_ocr: int = 1,
        fast_preprocess: bool = True,
        num_ctx: int = 4096,
        num_predict: int = 4096,
        # Обратная совместимость
        model_name: Optional[str] = None,
        vllm_url: Optional[str] = None,
        max_tokens: Optional[int] = None,
    ):
        self.pdf_path = pdf_path
        self.output_xlsx = output_xlsx
        self.poppler_path = poppler_path
        self.ollama_url = ollama_url.rstrip("/")
        self.ollama_model = model_name or ollama_model
        self.dpi = dpi
        self.preprocess_images = preprocess_images
        self.ocr_prompt = ocr_prompt or self.OCR_PROMPT
        self._df: Optional[pd.DataFrame] = None
        self.progress_callback = progress_callback or (lambda c, t, s: None)
        self.logger = print

        self.max_parallel_ocr = max(1, max_parallel_ocr)
        self.fast_preprocess = fast_preprocess
        self.num_ctx = num_ctx
        self.num_predict = max_tokens or num_predict
        self._log_lock = Lock()

    # ═══════════════════════════════════════════════════════════════
    # IMAGE PREPROCESSING
    # ═══════════════════════════════════════════════════════════════

    def _preprocess_image(self, img: Image.Image) -> Image.Image:
        if not self.preprocess_images:
            return img

        arr = np.array(img)
        if len(arr.shape) == 3 and arr.shape[2] == 4:
            arr = cv2.cvtColor(arr, cv2.COLOR_RGBA2RGB)
        if len(arr.shape) == 3:
            gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        else:
            gray = arr

        if not self.fast_preprocess:
            gray = cv2.bilateralFilter(gray, d=9, sigmaColor=75, sigmaSpace=75)

        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        return Image.fromarray(enhanced)

    def _image_to_base64(self, img: Image.Image) -> str:
        processed = self._preprocess_image(img)
        buf = BytesIO()
        processed.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("utf-8")

    # ═══════════════════════════════════════════════════════════════
    # OCR — через requests к Ollama REST API
    # ═══════════════════════════════════════════════════════════════

    def _ocr_from_image(self, img: Image.Image) -> str:
        """Отправляет изображение в Ollama через /api/chat и возвращает raw текст."""
        img_b64 = self._image_to_base64(img)

        payload = {
            "model": self.ollama_model,
            "messages": [
                {
                    "role": "user",
                    "content": self.ocr_prompt,
                    "images": [img_b64],
                }
            ],
            "options": {
                "temperature": 0.0,
                "num_ctx": self.num_ctx,
                "num_predict": self.num_predict,
            },
            "stream": False,
            "think": False,  # Отключаем thinking — экономит токены
        }

        # Отправляем запрос
        resp = requests.post(
            f"{self.ollama_url}/api/chat",
            json=payload,
            timeout=300,
        )
        resp.raise_for_status()

        # Парсим ответ
        data = resp.json()

        # DEBUG: логируем полный ответ (ключи)
        with self._log_lock:
            self.logger(f"  🔧 Ollama response keys: {list(data.keys())}")
            if "message" in data:
                msg = data["message"]
                self.logger(f"  🔧 message type: {type(msg)}, keys: {list(msg.keys()) if isinstance(msg, dict) else 'N/A'}")
                if isinstance(msg, dict):
                    content = msg.get("content", "")
                    self.logger(f"  🔧 content length: {len(content)}, first 1000: {repr(content[:1000])}")

        # Извлекаем текст
        raw = ""
        msg = data.get("message", {})
        if isinstance(msg, dict):
            raw = msg.get("content", "")
            # qwen3-vl кладёт ответ в thinking если content пустой
            if not raw and msg.get("thinking"):
                raw = msg["thinking"]
                with self._log_lock:
                    self.logger(f"  🔧 Взяли текст из 'thinking' ({len(raw)} символов)")
        elif isinstance(msg, str):
            raw = msg
        elif "response" in data:
            raw = data["response"]

        raw = raw.strip()

        # Вырезаем <think>...</think> теги
        raw = re.sub(r'<think>.*?</think>', '', raw, flags=re.DOTALL).strip()

        # Если текст начинается с рассуждений — извлекаем только таблицу
        # Ищем начало markdown таблицы (первая строка с |)
        lines = raw.split("\n")
        table_start = -1
        for i, line in enumerate(lines):
            if line.strip().startswith("|"):
                table_start = i
                break

        if table_start > 0:
            # Есть рассуждения перед таблицей — отрезаем их
            raw = "\n".join(lines[table_start:]).strip()
            with self._log_lock:
                self.logger(f"  🔧 Обрезали рассуждения, таблица с строки {table_start}")

        # Также проверяем HTML таблицу
        if "<table" in raw.lower():
            html_start = raw.lower().find("<table")
            if html_start > 0:
                raw = raw[html_start:].strip()

        # Убираем markdown обёртку
        raw = re.sub(r'^```(?:markdown)?\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
        raw = raw.strip()

        return raw

    def _detect_markdown_sections(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Детект разделов ТОЛЬКО по разметке от модели.

        Заголовок: **Текст** + пустой № п/п
        Подзаголовок: *Текст* + заполнен № п/п

        Никаких эвристик по тексту или колонкам — только разметка.
        """
        if df.empty:
            return df

        df = df.copy()
        df['_is_section'] = False
        df['_header'] = ''
        df['_subheader'] = ''

        section_count = 0
        header_count = 0
        subheader_count = 0

        for idx in df.index:
            name = str(df.at[idx, 'Наименование']).strip()
            num = str(df.at[idx, '№ п/п']).strip()
            unit = str(df.at[idx, 'Ед. изм.']).strip()
            qty = str(df.at[idx, 'Кол-во']).strip()

            has_data = bool(unit) or bool(qty)
            has_num = bool(num) and re.match(r'^\d', num)

            # 🔥 Строки с данными — никогда не разделы (даже с разметкой)
            if has_data:
                continue

            # ═══════════════════════════════════════════════════════════════
            # ЗАГОЛОВОК: **Текст** + пустой № п/п
            # ═══════════════════════════════════════════════════════════════

            if not has_num and re.match(r'^\*\*.+\*\*$', name):
                df.at[idx, '_is_section'] = True
                section_count += 1
                clean_name = re.sub(r'\*\*(.+)\*\*', r'\1', name).strip()
                df.at[idx, 'Наименование'] = clean_name
                df.at[idx, '_header'] = clean_name
                df.at[idx, '_subheader'] = ''
                header_count += 1
                continue

            # ═══════════════════════════════════════════════════════════════
            # ПОДЗАГОЛОВОК: *Текст* + заполнен № п/п
            # ═══════════════════════════════════════════════════════════════

            if has_num and re.match(r'^\*[^*].*\*$', name):
                df.at[idx, '_is_section'] = True
                section_count += 1
                clean_name = re.sub(r'\*(.+)\*', r'\1', name).strip()
                df.at[idx, 'Наименование'] = clean_name
                df.at[idx, '_subheader'] = clean_name
                subheader_count += 1
                continue

        self.logger(f"🔖 Найдено разделов: {section_count} (заголовки: {header_count}, подзаголовки: {subheader_count})")
        return df

    def _ocr_page(self, page: Image.Image) -> pd.DataFrame:
        raw = self._ocr_from_image(page)

        if not raw:
            raise ValueError("Empty response from Ollama")

        with self._log_lock:
            self.logger(f"🔍 OCR raw (first 500 chars): {repr(raw[:500])}")

        df = self._parse_ocr_output(raw)

        # 🔥 НОВОЕ: Ищем разделы по **жирному**
        df = self._detect_markdown_sections(df)

        return df

    # ═══════════════════════════════════════════════════════════════
    # PARSING — универсальный (markdown, HTML, plain text)
    # ═══════════════════════════════════════════════════════════════

    def _parse_ocr_output(self, text: str) -> pd.DataFrame:
        lines = text.strip().splitlines()

        # HTML?
        if "<table" in text.lower() or "<tr" in text.lower():
            return self._parse_html_table(text)

        # Markdown?
        has_pipes = sum(1 for l in lines if "|" in l) > len(lines) * 0.3
        if has_pipes:
            return self._parse_markdown_lines(lines)

        # Plain text
        return self._parse_plain_text(lines)

    def _parse_html_table(self, html: str) -> pd.DataFrame:
        dfs = pd.read_html(StringIO(html))
        if not dfs:
            raise ValueError("No HTML tables found")
        df = dfs[0].fillna("").astype(str)
        df.columns = [str(c).strip() for c in df.columns]
        return self._normalize_columns(df)

    def _parse_markdown_lines(self, lines: list) -> pd.DataFrame:
        table_lines = []
        for line in lines:
            line = line.strip()
            if not line.startswith("|"):
                continue
            if not line.endswith("|"):
                line += "|"
            inner = re.sub(r"^\||\|$", "", line).strip()
            if re.match(r"^[\s\-:|]+$", inner):
                continue
            table_lines.append(line)

        if len(table_lines) < 2:
            raise ValueError(f"Too few markdown lines ({len(table_lines)})")

        header_cells = [c.strip() for c in table_lines[0].strip("|").split("|")]
        n_cols = len(header_cells)

        rows = []
        for line in table_lines[1:]:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) > n_cols:
                cells = cells[:n_cols]
            while len(cells) < n_cols:
                cells.append("")
            rows.append(cells)

        df = pd.DataFrame(rows, columns=header_cells)
        return self._normalize_columns(df)

    def _parse_plain_text(self, lines: list) -> pd.DataFrame:
        if not lines:
            raise ValueError("Empty OCR output")

        data_lines = lines[1:] if lines else []
        rows = []
        num_pattern = re.compile(r'^(\d+[\d.]*)\s+(.+)')
        unit_qty_pattern = re.compile(
            r'^(.*?)\s+(Т|т|м|М|м2|м3|мм|шт|кг|кГ|л|км|исп\.?|компл\.?|п\.м\.?)\s+([\d.,]+)\s*$'
        )

        for line in data_lines:
            line = line.strip()
            if not line:
                continue

            m_num = num_pattern.match(line)
            if m_num:
                num = m_num.group(1)
                rest = m_num.group(2).strip()
                m_uq = unit_qty_pattern.match(rest)
                if m_uq:
                    rows.append({"№ п/п": num, "Наименование": m_uq.group(1).strip(),
                                 "Ед. изм.": m_uq.group(2).strip(), "Кол-во": m_uq.group(3).strip()})
                else:
                    rows.append({"№ п/п": num, "Наименование": rest, "Ед. изм.": "", "Кол-во": ""})
            else:
                rows.append({"№ п/п": "", "Наименование": line, "Ед. изм.": "", "Кол-во": ""})

        if not rows:
            raise ValueError("No data rows parsed")

        df = pd.DataFrame(rows)
        df = df[~df.apply(lambda row: all(str(v).strip() == "" for v in row), axis=1)].reset_index(drop=True)
        return df

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        col_mapping = self._detect_columns(df.columns.tolist())
        rename_map = {}
        for std_name, real_name in col_mapping.items():
            if real_name and real_name in df.columns:
                rename_map[real_name] = std_name
        df = df.rename(columns=rename_map)

        expected = ["№ п/п", "Наименование", "Ед. изм.", "Кол-во"]
        for col in expected:
            if col not in df.columns:
                df[col] = ""
        df = df[expected].copy()

        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()

        df = df[~df.apply(lambda row: all(str(v).strip() == "" for v in row), axis=1)].reset_index(drop=True)
        return df

    def _detect_columns(self, columns: List[str]) -> dict:
        mapping = {"№ п/п": None, "Наименование": None, "Ед. изм.": None, "Кол-во": None}
        cols_lower = {c: c.strip().lower() for c in columns}

        for real_name, lower_name in cols_lower.items():
            if any(x in lower_name for x in ["№", "п/п", "строк", "номер"]):
                mapping["№ п/п"] = real_name
            elif any(x in lower_name for x in ["наименование", "вид работ", "название"]):
                mapping["Наименование"] = real_name
            elif any(x in lower_name for x in ["ед.", "изм", "единиц"]):
                mapping["Ед. изм."] = real_name
            elif any(x in lower_name for x in ["кол", "колич", "объём", "объем"]):
                mapping["Кол-во"] = real_name

        return mapping

    # ═══════════════════════════════════════════════════════════════
    # POST-PROCESSING
    # ═══════════════════════════════════════════════════════════════

    def _merge_multiline_names(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        rows = df.to_dict("records")
        merged = []
        i = 0

        while i < len(rows):
            cur = rows[i]

            # 🔥 Если это раздел (помечен ранее) — не склеиваем
            if cur.get("_is_section", False):
                merged.append(cur.copy())
                i += 1
                continue

            cur_num = str(cur.get("№ п/п", "")).strip()
            cur_name = str(cur.get("Наименование", "")).strip()

            if cur_num:
                combined_name = cur_name
                j = i + 1
                while j < len(rows):
                    nxt = rows[j]
                    nxt_num = str(nxt.get("№ п/п", "")).strip()
                    nxt_name = str(nxt.get("Наименование", "")).strip()

                    if not nxt_num and nxt_name:
                        if re.match(r'^\d+\s+[А-ЯЁA-Z]', nxt_name):
                            break
                        combined_name += " " + nxt_name
                        j += 1
                    else:
                        break

                new_row = cur.copy()
                new_row["Наименование"] = combined_name.strip()
                for k in range(i + 1, j):
                    nxt = rows[k]
                    if not new_row.get("Ед. изм.") and nxt.get("Ед. изм."):
                        new_row["Ед. изм."] = nxt["Ед. изм."]
                    if not new_row.get("Кол-во") and nxt.get("Кол-во"):
                        new_row["Кол-во"] = nxt["Кол-во"]
                merged.append(new_row)
                i = j
            else:
                merged.append(cur)
                i += 1

        return pd.DataFrame(merged)

    def _merge_multiline_subheaders(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Склеивает подзаголовки, которые перенесены на несколько строк.
        Работает ТОЛЬКО со строками, помеченными как _subheader.
        """
        if df.empty:
            return df

        df = df.copy()
        rows = df.to_dict("records")
        merged = []
        i = 0

        while i < len(rows):
            cur = rows[i]
            cur_is_section = cur.get("_is_section", False)
            cur_subheader = cur.get("_subheader", "")
            cur_name = str(cur.get("Наименование", "")).strip()
            cur_has_subheader = bool(cur_subheader)

            # 🔥 Если это подзаголовок (помечен ранее)
            if cur_is_section and cur_has_subheader:
                combined_subheader = cur_name
                j = i + 1

                while j < len(rows):
                    nxt = rows[j]
                    nxt_name = str(nxt.get("Наименование", "")).strip()
                    nxt_num = str(nxt.get("№ п/п", "")).strip()
                    nxt_unit = str(nxt.get("Ед. изм.", "")).strip()
                    nxt_qty = str(nxt.get("Кол-во", "")).strip()
                    nxt_has_data = bool(nxt_unit) or bool(nxt_qty)
                    nxt_is_section = nxt.get("_is_section", False)

                    # 🔥 Критерии продолжения:
                    # 1. Не раздел
                    # 2. Нет данных
                    # 3. Начинается с маленькой буквы ИЛИ тире
                    if not nxt_is_section and not nxt_has_data and nxt_name:
                        if nxt_name[0].islower() or nxt_name.startswith((' ', '-', '–', '—')):
                            combined_subheader += " " + nxt_name
                            j += 1
                        else:
                            break
                    else:
                        break

                new_row = cur.copy()
                new_row["Наименование"] = combined_subheader.strip()
                new_row["_subheader"] = combined_subheader.strip()
                merged.append(new_row)
                i = j
            else:
                merged.append(cur.copy())
                i += 1

        result_df = pd.DataFrame(merged)
        self.logger(f"🔗 Склеено подзаголовков: {len(rows) - len(result_df)} строк")
        return result_df

    def _build_hierarchy(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        parents, levels, paths = [], [], []
        for val in df["№ п/п"]:
            s = str(val).strip()
            if not s:
                parents.append("")
                levels.append(0)
                paths.append("")
                continue

            clean = re.sub(r"[^\d.]", "", s)
            parts = [p for p in clean.split(".") if p.isdigit()]
            level = len(parts)
            parent = ".".join(parts[:-1]) if level > 1 else ""
            path = " > ".join(".".join(parts[:i+1]) for i in range(len(parts))) if parts else ""
            parents.append(parent)
            levels.append(level)
            paths.append(path)

        df = df.copy()
        df["parent_id"] = parents
        df["level"] = levels
        df["path"] = paths
        return df

    def _propagate_headers_across_pages(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Пропагирует заголовки и подзаголовки по всему DataFrame.
        """
        if df.empty:
            return df

        df = df.copy()
        current_header = ''
        current_subheader = ''

        for idx in df.index:
            is_section = df.at[idx, '_is_section']
            header_val = df.at[idx, '_header']
            subheader_val = df.at[idx, '_subheader']

            if is_section:
                if header_val:  # Заголовок
                    current_header = header_val
                    current_subheader = ''
                elif subheader_val:  # Подзаголовок
                    current_subheader = subheader_val
            else:
                # Обычная строка — пропагируем
                df.at[idx, '_header'] = current_header
                df.at[idx, '_subheader'] = current_subheader

        self.logger(f"🔗 Пропагация заголовков завершена")
        return df

    def _remove_before_first_header(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Удаляет все строки до первого ЗАГОЛОВКА (**...**).
        """
        if df.empty:
            return df

        df = df.copy()

        # Ищем первую строку с _header
        first_header_idx = None
        for idx in df.index:
            if df.at[idx, '_header']:  # Заголовок найден
                first_header_idx = idx
                break

        if first_header_idx is not None:
            rows_before = df.index.get_loc(first_header_idx)
            if rows_before > 0:
                df = df.iloc[first_header_idx:].reset_index(drop=True)
                self.logger(f"🗑️ Удалено {rows_before} строк до первого заголовка")

        return df

    # ═══════════════════════════════════════════════════════════════
    # EXCEL
    # ═══════════════════════════════════════════════════════════════

    def _save_to_excel(self, df: pd.DataFrame, path: str):
        from openpyxl.styles import Font, PatternFill, Alignment

        if df.empty:
            df = pd.DataFrame([{"№ п/п": "НЕТ ДАННЫХ", "Наименование": "", "Ед. изм.": "", "Кол-во": ""}])

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            cols_order = ["№ п/п", "Раздел", "Подраздел", "Наименование", "Ед. изм.", "Кол-во"]

            final_df = pd.DataFrame()
            for col in cols_order:
                if col == 'Раздел':
                    final_df[col] = df.get('_header', '')
                elif col == 'Подраздел':
                    final_df[col] = df.get('_subheader', '')
                elif col in df.columns:
                    final_df[col] = df[col]
                else:
                    final_df[col] = ''

            final_df.to_excel(writer, index=False, sheet_name="Ведомость")
            ws = writer.book["Ведомость"]
            ws.freeze_panes = "A2"

            # Настройка ширины колонок
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10

            # Выделение заголовков жирным
            bold_font = Font(bold=True)
            gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            for row_idx, df_idx in enumerate(df.index, start=2):
                # 🔥 Выделяем только заголовки (по флагу _header)
                if df.at[df_idx, '_header']:
                    for col_idx in range(1, len(cols_order) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = bold_font
                        cell.fill = gray_fill

        self.logger(f"✅ Сохранено: {path}")

    # ═══════════════════════════════════════════════════════════════
    # CORE
    # ═══════════════════════════════════════════════════════════════

    def _process_single_page(self, page_idx: int, page: Image.Image, total_pages: int) -> Optional[pd.DataFrame]:
        try:
            with self._log_lock:
                self.logger(f"📄 Страница {page_idx + 1}/{total_pages} — OCR...")

            # 1. Получаем сырой текст от модели
            raw = self._ocr_from_image(page)

            if not raw:
                raise ValueError("Empty response from Ollama")

            # 2. Парсим в DataFrame
            df_page = self._parse_ocr_output(raw)

            # 3. Находим разделы (только маркировка, без пропагации)
            df_page = self._detect_markdown_sections(df_page)

            # 4. Добавляем номер страницы для трассировки
            df_page["page"] = page_idx + 1

            with self._log_lock:
                self.logger(f"  ✅ Страница {page_idx + 1}: {len(df_page)} строк")

            return df_page

        except Exception as e:
            with self._log_lock:
                self.logger(f"  ⚠️ Ошибка на странице {page_idx + 1}: {e}")
            return None

    def process(self) -> pd.DataFrame:
        self.logger(f"📄 PDF: {self.pdf_path} (DPI={self.dpi})")
        self.logger(f"⚡ Ollama: {self.ollama_url}, model={self.ollama_model}, parallel={self.max_parallel_ocr}")

        try:
            pages = convert_from_path(self.pdf_path, dpi=self.dpi, poppler_path=self.poppler_path)
        except Exception as e:
            raise RuntimeError(f"pdf2image failed: {e}")

        total = len(pages)
        self.logger(f"📄 Страниц: {total}")

        all_dfs = []

        if self.max_parallel_ocr <= 1:
            # Последовательная обработка
            for i, page in enumerate(pages):
                self.progress_callback(i + 1, total, "📄 OCR")
                df_page = self._process_single_page(i, page, total)
                if df_page is not None:
                    all_dfs.append(df_page)
        else:
            # Параллельная обработка
            all_dfs_dict = {}
            completed = 0
            lock = Lock()

            with ThreadPoolExecutor(max_workers=self.max_parallel_ocr) as executor:
                futures = {executor.submit(self._process_single_page, i, page, total): i
                           for i, page in enumerate(pages)}
                for future in as_completed(futures):
                    idx = futures[future]
                    try:
                        df_page = future.result()
                        if df_page is not None:
                            with lock:
                                all_dfs_dict[idx] = df_page
                    except Exception as e:
                        self.logger(f"  💥 Страница {idx + 1}: {e}")
                    with lock:
                        completed += 1
                    self.progress_callback(completed, total, "📄 OCR")

            all_dfs = [all_dfs_dict[i] for i in sorted(all_dfs_dict.keys())]

            self.logger(f"✅ OCR завершён: собрано {len(all_dfs)} страниц")

        # Финализация
        if not all_dfs:
            self._df = pd.DataFrame()
        else:
            self.logger(f"🔧 Начало структуризации...")

            # 1. Объединяем все страницы в один DataFrame
            self._df = pd.concat(all_dfs, ignore_index=True)
            self.logger(f"   📊 Объединено: {len(self._df)} строк")

            # 2. Слияние многострочных наименований (работы/материалы)
            self._df = self._merge_multiline_names(self._df)
            self.logger(f"   🔗 Слияние работ завершено: {len(self._df)} строк")

            # 3. 🔥 НОВОЕ: Слияние многострочных подзаголовков
            self._df = self._merge_multiline_subheaders(self._df)

            # 4. Пропагация заголовков по всему документу
            self._df = self._propagate_headers_across_pages(self._df)

            # 5. Построение иерархии (level, parent_id, path)
            self._df = self._build_hierarchy(self._df)

            # 6. Удаление строк до первого заголовка
            self._df = self._remove_before_first_header(self._df)

            self.logger(f"✅ Структуризация завершена: {len(self._df)} строк")

        # Выгрузка модели
        try:
            requests.post(f"{self.ollama_url}/api/generate",
                          json={"model": self.ollama_model, "prompt": "", "keep_alive": 0}, timeout=10)
            self.logger(f"🧹 {self.ollama_model} выгружена из VRAM")
        except:
            pass

        self.logger(f"✅ Итого: {len(self._df)} строк")
        return self._df

    # ═══════════════════════════════════════════════════════════════
    # PUBLIC API
    # ═══════════════════════════════════════════════════════════════

    def get_dataframe(self) -> Optional[pd.DataFrame]:
        return self._df.copy() if self._df is not None else None

    def save_to_excel(self, output_path: Optional[str] = None):
        path = output_path or self.output_xlsx
        if not path:
            raise ValueError("output_path не задан")
        if self._df is None:
            raise RuntimeError("DataFrame не создан. Вызовите .process()")
        self._save_to_excel(self._df, path)