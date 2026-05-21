import pandas as pd
from pathlib import Path
from typing import Union


class FinalAggregator:
    PARSE_EXTRA_COLUMNS = [
        'Наименование ВиКР документа',
        'Марка РД',
        'Номер проекта',
        'Наименование зданий, сооружений, систем и установок',
        'structure',
    ]

    def __init__(self, progress_callback=None, log_callback=None):
        self.progress_callback = progress_callback or (lambda c, t, s: None)
        self._log_callback = log_callback

    @staticmethod
    def _normalize_unit(series: pd.Series) -> pd.Series:
        return (
            series.fillna('')
            .astype(str)
            .str.strip()
            .str.replace(r'[.,\s]', '', regex=True)
            .str.lower()
        )

    @staticmethod
    def _extract_vikr_unit(vikr_text) -> str:
        if not isinstance(vikr_text, str) or '|' not in vikr_text:
            return ''
        parts = vikr_text.split('|')
        return parts[1].strip() if len(parts) >= 2 else ''

    @staticmethod
    def _fmt_num(v):
        if pd.isna(v):
            return ''
        if isinstance(v, float):
            return str(int(v)) if v.is_integer() else str(v)
        if isinstance(v, int):
            return str(v)
        s = str(v).strip()
        if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
            return s[:-2]
        return s

    @staticmethod
    def _clean_note(note) -> str:
        if note is None or pd.isna(note):
            return ''
        return str(note).strip()

    @classmethod
    def _merge_notes(cls, *notes) -> str:
        merged = []
        seen = set()
        for note in notes:
            raw = cls._clean_note(note)
            if not raw:
                continue
            for part in [p.strip() for p in raw.split(';') if p.strip()]:
                key = part.lower()
                if key not in seen:
                    seen.add(key)
                    merged.append(part)
        return '; '.join(merged)

    @classmethod
    def _note_tags(cls, note) -> set:
        text = cls._clean_note(note).lower()
        tags = set()
        if 'работа не выбрана' in text:
            tags.add('no_answer')
        if 'низкая уверенность модели' in text:
            tags.add('low_confidence')
        if 'единица измерения не совпадает' in text or 'не совпадает единица измерения' in text:
            tags.add('unit_mismatch')
        return tags

    @classmethod
    def _pick_fill_color(cls, note):
        tags = cls._note_tags(note)
        if len(tags) >= 2:
            return 'FFC4293E'
        if 'low_confidence' in tags:
            return 'FFFFA500'
        if 'unit_mismatch' in tags:
            return 'FFF4CCCC'
        if 'no_answer' in tags:
            return 'FFFFFF00'
        return None

    def aggregate(self, input_path: Union[str, Path]) -> pd.DataFrame:
        input_path = Path(input_path)
        self._log(f"📊 Загрузка файла: {input_path.name}")

        df = pd.read_excel(input_path, dtype={'Код работы': str})

        if 'Наименование ВиКР' in df.columns:
            df = df.rename(columns={'Наименование ВиКР': 'Наименование ВиКР документа'})

        required_cols = ['Наименование ВР', 'Кол-во', 'Код работы', 'Смэтченная работа ВиКР']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {missing}")

        self._log("🧹 Очистка данных...")

        df['Кол-во'] = pd.to_numeric(df['Кол-во'], errors='coerce').fillna(0)
        for col in ['Раздел', 'Подраздел']:
            if col not in df.columns:
                df[col] = ''
            df[col] = df[col].fillna('').astype(str).str.strip()

        df['Наименование ВР'] = df['Наименование ВР'].fillna('').astype(str).str.strip()
        if 'Ед. изм.' not in df.columns:
            df['Ед. изм.'] = ''
        df['Ед. изм.'] = df['Ед. изм.'].fillna('').astype(str).str.strip()
        df['Смэтченная работа ВиКР'] = df['Смэтченная работа ВиКР'].fillna('').astype(str).str.strip()
        for col in self.PARSE_EXTRA_COLUMNS:
            if col not in df.columns:
                df[col] = ''
            df[col] = df[col].fillna('').astype(str).str.strip()
        df['Наименование работ'] = df['Наименование работ'].where(
            df['Наименование работ'].astype(bool),
            df['Наименование ВР'],
        )
        if 'Ошибки' not in df.columns:
            df['Ошибки'] = ''
        df['Ошибки'] = df['Ошибки'].apply(self._clean_note)

        if '№ п/п' in df.columns:
            df['№ п/п'] = df['№ п/п'].apply(self._fmt_num)
        else:
            df['№ п/п'] = ''

        if 'Тип' in df.columns:
            df = df[df['Тип'].fillna('').astype(str).str.strip() == 'work'].copy()

        df['Код работы'] = df['Код работы'].fillna('').astype(str).str.strip()
        df['_orig_index'] = df.index

        if df.empty:
            self._log("⚠️ Нет строк для обработки после очистки")
            return pd.DataFrame()

        self._log(f"   Строк после очистки: {len(df)}")

        matched_text = df['Смэтченная работа ВиКР'].fillna('').astype(str).str.strip()
        passthrough_mask = (matched_text == '') | (df['Код работы'] == '')
        no_match_df = df[passthrough_mask].copy()
        df_matched = df[~passthrough_mask].copy()

        self._log(f"   Без выбора/кода (идут без агрегации): {len(no_match_df)}")
        self._log(f"   С корректным мэтчем: {len(df_matched)}")

        df_matched['_vikr_unit'] = df_matched['Смэтченная работа ВиКР'].apply(self._extract_vikr_unit)

        vr_unit_norm = self._normalize_unit(df_matched['Ед. изм.'])
        vikr_unit_norm = self._normalize_unit(df_matched['_vikr_unit'])

        unit_mismatch_mask = (
            (vikr_unit_norm != '') &
            (vr_unit_norm != '') &
            (vr_unit_norm != vikr_unit_norm)
        )
        unit_mismatch_df = df_matched[unit_mismatch_mask].copy()
        df_to_agg = df_matched[~unit_mismatch_mask].copy()

        self._log(f"   Несовпадение ед. изм. (идут без агрегации): {len(unit_mismatch_df)}")
        self._log(f"   Для агрегации: {len(df_to_agg)}")

        no_match_df['_vikr_name'] = no_match_df['Наименование ВР']
        no_match_df['_note'] = no_match_df['Ошибки'].apply(self._clean_note)

        unit_mismatch_df['_vikr_name'] = unit_mismatch_df['Наименование ВР']
        unit_mismatch_df['_note'] = unit_mismatch_df.apply(
            lambda row: self._merge_notes(row.get('Ошибки', ''), 'Единица измерения не совпадает'),
            axis=1,
        )

        df_to_agg['_vikr_name'] = df_to_agg['Смэтченная работа ВиКР'].apply(
            lambda x: x.split('|')[0].strip() if isinstance(x, str) and '|' in x else str(x)
        )
        df_to_agg['_note'] = df_to_agg['Ошибки'].apply(self._clean_note)

        for col in ['Уверенность модели', 'Score эмбеддинга', 'Позиция в топ-5']:
            if col in df_to_agg.columns:
                df_to_agg[col] = pd.to_numeric(df_to_agg[col], errors='coerce').fillna(0)

        def _agg_qty_block(series: pd.Series) -> float:
            vals = series.dropna().tolist()
            if not vals:
                return 0.0
            if len(vals) == 1:
                return vals[0]
            if len(set(vals)) == 1:
                return vals[0]

            max_val = max(vals)
            others_sum = sum(vals) - max_val
            if abs(others_sum - max_val) < 1e-6:
                return max_val

            return sum(vals)

        self._log("🧮 Агрегация данных...")
        self.progress_callback(1, 3, "📊 Агрегация")

        if not df_to_agg.empty:
            df_to_agg['group_key'] = (
                df_to_agg['Раздел'].astype(str) + '||' +
                df_to_agg['Подраздел'].astype(str) + '||' +
                df_to_agg['Код работы'].astype(str)
            )

            df_to_agg = df_to_agg.sort_values('_orig_index').reset_index(drop=True)

            changed_name = df_to_agg['_vikr_name'] != df_to_agg['_vikr_name'].shift(1)
            changed_group = df_to_agg['group_key'] != df_to_agg['group_key'].shift(1)
            is_new_block = changed_name | changed_group
            df_to_agg['block_id'] = is_new_block.cumsum()
            df_to_agg['seq_key'] = df_to_agg['group_key'] + '_blk_' + df_to_agg['block_id'].astype(str)

            seq_agg_dict = {
                'Кол-во': _agg_qty_block,
                '№ п/п': lambda x: ', '.join(v for v in x.unique() if v),
                '_vikr_name': 'first',
                'Смэтченная работа ВиКР': 'first',
                'Код работы': 'first',
                'Раздел': 'first',
                'Подраздел': 'first',
                'Наименование ВР': lambda x: ' | '.join(x.unique()),
                '_orig_index': 'first',
                'group_key': 'first',
                'seq_key': 'first',
                '_note': lambda x: self._merge_notes(*x.tolist()),
            }
            if 'Ед. изм.' in df_to_agg.columns:
                seq_agg_dict['Ед. изм.'] = lambda x: next((v for v in x if str(v).strip()), '')
            if 'Уверенность модели' in df_to_agg.columns:
                seq_agg_dict['Уверенность модели'] = 'mean'
            if 'Score эмбеддинга' in df_to_agg.columns:
                seq_agg_dict['Score эмбеддинга'] = 'mean'
            if 'Позиция в топ-5' in df_to_agg.columns:
                seq_agg_dict['Позиция в топ-5'] = 'first'
            for col in self.PARSE_EXTRA_COLUMNS:
                if col in df_to_agg.columns:
                    if col == 'Наименование работ':
                        seq_agg_dict[col] = lambda x: ' | '.join(v for v in x.unique() if v)
                    else:
                        seq_agg_dict[col] = 'first'

            seq_agg = df_to_agg.groupby('seq_key', sort=False).agg(seq_agg_dict).reset_index(drop=True)

            seq_sizes = df_to_agg.groupby('seq_key', sort=False).size().reset_index(
                name='Количество объединённых работ'
            )
            seq_agg = seq_agg.merge(seq_sizes, on='seq_key', how='left')

            final_agg_dict = {
                'Кол-во': 'sum',
                '№ п/п': lambda x: ', '.join(v for v in x.unique() if v),
                'Наименование ВР': lambda x: ' | '.join(x.unique()),
                '_vikr_name': 'first',
                'Смэтченная работа ВиКР': 'first',
                'Код работы': 'first',
                'Раздел': 'first',
                'Подраздел': 'first',
                '_orig_index': 'first',
                'group_key': 'first',
                '_note': lambda x: self._merge_notes(*x.tolist()),
            }
            if 'Ед. изм.' in seq_agg.columns:
                final_agg_dict['Ед. изм.'] = 'first'
            if 'Уверенность модели' in seq_agg.columns:
                final_agg_dict['Уверенность модели'] = 'mean'
            if 'Score эмбеддинга' in seq_agg.columns:
                final_agg_dict['Score эмбеддинга'] = 'mean'
            if 'Позиция в топ-5' in seq_agg.columns:
                final_agg_dict['Позиция в топ-5'] = 'first'
            for col in self.PARSE_EXTRA_COLUMNS:
                if col in seq_agg.columns:
                    if col == 'Наименование работ':
                        final_agg_dict[col] = lambda x: ' | '.join(v for v in x.unique() if v)
                    else:
                        final_agg_dict[col] = 'first'

            aggregated_df = seq_agg.groupby('group_key', sort=False).agg(final_agg_dict).reset_index(drop=True)
            counts = seq_agg.groupby('group_key', sort=False)['Количество объединённых работ'].sum()
            aggregated_df['Количество объединённых работ'] = aggregated_df['group_key'].map(counts)
        else:
            aggregated_df = pd.DataFrame()

        self.progress_callback(2, 3, "📊 Агрегация")

        passthrough_parts = []
        for part_df in [no_match_df, unit_mismatch_df]:
            if not part_df.empty:
                pt = part_df.copy()
                pt['Количество объединённых работ'] = 1
                passthrough_parts.append(pt)

        self._log("📋 Формирование итоговой таблицы...")
        parts_to_concat = [aggregated_df] if not aggregated_df.empty else []
        parts_to_concat.extend(passthrough_parts)

        if not parts_to_concat:
            self._log("⚠️ Нет данных для формирования результата")
            return pd.DataFrame()

        final_df = pd.concat(parts_to_concat, ignore_index=True)

        if '_orig_index' in final_df.columns:
            final_df = final_df.sort_values('_orig_index').reset_index(drop=True)
            final_df = final_df.drop(columns=['_orig_index'])

        for col in ['_vikr_unit', 'group_key', 'block_id', 'seq_key']:
            if col in final_df.columns:
                final_df = final_df.drop(columns=[col])

        final_df = final_df.rename(columns={
            '_vikr_name': 'Наименование ВиКР',
            'Наименование ВР': 'Наименования работ ВОР',
            '№ п/п': 'Объединённые № п/п',
            '_note': 'Примечание',
        })

        if 'Тип' in final_df.columns:
            final_df = final_df.drop(columns=['Тип'])

        col_order = [
            'Объединённые № п/п',
            'Наименование ВиКР',
            'Ед. изм.',
            'Кол-во',
            'Раздел',
            'Подраздел',
            'Код работы',
            'Смэтченная работа ВиКР',
            'Позиция в топ-5',
            'Уверенность модели',
            'Score эмбеддинга',
            'Наименования работ ВОР',
            'Количество объединённых работ',
            'Примечание',
            *self.PARSE_EXTRA_COLUMNS,
        ]
        col_order = [c for c in col_order if c in final_df.columns]
        extra_cols = [c for c in final_df.columns if c not in col_order]
        final_df = final_df[col_order + extra_cols]

        self.progress_callback(3, 3, "📊 Агрегация")
        self._log(f"✅ Агрегация завершена: {len(final_df)} строк (из них агрегировано: {len(aggregated_df)})")
        return final_df

    @staticmethod
    def save_df_to_excel_with_highlights(df: pd.DataFrame, output_path: Union[str, Path]):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        output_path = Path(output_path)
        wb = Workbook()
        ws = wb.active

        headers = list(df.columns)
        ws.append(headers)

        qty_col_idx = headers.index('Кол-во') + 1 if 'Кол-во' in headers else None

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            values = [None if pd.isna(row[col]) else row[col] for col in headers]
            ws.append(values)

            if qty_col_idx:
                cell = ws.cell(row=row_idx, column=qty_col_idx)
                try:
                    val = float(row['Кол-во'])
                    if val == int(val):
                        cell.value = int(val)
                        cell.number_format = '0'
                    else:
                        cell.value = val
                        cell.number_format = '0.######'
                except (ValueError, TypeError):
                    pass

            row_note = FinalAggregator._merge_notes(
                row.get('Ошибки', ''),
                row.get('Примечание', ''),
            )
            fill_color = FinalAggregator._pick_fill_color(row_note)
            if fill_color:
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill

        for col_idx, header in enumerate(headers, 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(max(len(str(header)) + 2, 10), 50)

        wb.save(str(output_path))

    def save_to_excel(self, df: pd.DataFrame, output_path: Union[str, Path]):
        self.save_df_to_excel_with_highlights(df, output_path)
        self._log(f"💾 Результат сохранён: {Path(output_path).name}")

    def _log(self, msg: str):
        if self._log_callback:
            self._log_callback(msg)
        else:
            print(msg)
