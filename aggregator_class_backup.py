# aggregator_class.py — OPTIMIZED v3.3
# Изменения v3.3:
#   1. Строки без матча НЕ выкидываются, а идут в исходном порядке
#   2. Для них в колонке "Наименование ВиКР" пишется "Не найдена подходящая"
#   3. Логика полностью совпадает с обработкой несовпавших ед. изм.
#   4. Сохранён исходный порядок всех строк через _orig_index

import pandas as pd
from pathlib import Path
from typing import Union


class FinalAggregator:
    """
    Финальная структуризация: объединение работ по кодам ВиКР с умным суммированием количеств.
    Сохраняет исходный порядок строк. Работы без матча или с несовпавшими ед. изм.
    проходят без агрегации, но остаются в общей таблице.
    """

    def __init__(self, progress_callback=None, log_callback=None):
        self.progress_callback = progress_callback or (lambda c, t, s: None)
        self._log_callback = log_callback

    @staticmethod
    def _normalize_unit(series: pd.Series) -> pd.Series:
        """Нормализует ед. изм.: убирает точки, запятые, пробелы, приводит к нижнему регистру."""
        return (
            series.fillna('')
            .astype(str)
            .str.strip()
            .str.replace(r'[.,\s]', '', regex=True)
            .str.lower()
        )

    @staticmethod
    def _extract_vikr_unit(vikr_text) -> str:
        """Извлекает ед. изм. из 'Смэтченная работа ВиКР' (формат: 'name | unit | composition')."""
        if not isinstance(vikr_text, str) or '|' not in vikr_text:
            return ''
        parts = vikr_text.split('|')
        return parts[1].strip() if len(parts) >= 2 else ''

    @staticmethod
    def _fmt_num(v):
        """Форматирует № п/п: 1.0 → '1'."""
        if pd.isna(v): return ''
        if isinstance(v, float): return str(int(v)) if v.is_integer() else str(v)
        if isinstance(v, int): return str(v)
        s = str(v).strip()
        if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
            return s[:-2]
        return s

    def aggregate(self, input_path: Union[str, Path]) -> pd.DataFrame:
        input_path = Path(input_path)
        self._log(f"📊 Загрузка файла: {input_path.name}")

        df = pd.read_excel(input_path, dtype={'Код работы': str})

        required_cols = ['Наименование ВР', 'Кол-во', 'Код работы', 'Смэтченная работа ВиКР']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {missing}")

        self._log("🧹 Очистка данных...")

        df['Кол-во'] = pd.to_numeric(df['Кол-во'], errors='coerce').fillna(0)
        for col in ['Раздел', 'Подраздел']:
            if col not in df.columns: df[col] = ''
            df[col] = df[col].fillna('').astype(str).str.strip()

        df['Наименование ВР'] = df['Наименование ВР'].fillna('').astype(str).str.strip()
        df['Ед. изм.'] = df['Ед. изм.'].fillna('').astype(str).str.strip() if 'Ед. изм.' in df.columns else ''
        df['Смэтченная работа ВиКР'] = df['Смэтченная работа ВиКР'].fillna('').astype(str).str.strip()

        if '№ п/п' in df.columns:
            df['№ п/п'] = df['№ п/п'].apply(self._fmt_num)
        else:
            df['№ п/п'] = ''

        # 🔑 Сохраняем исходный индекс для восстановления порядка в конце
        df['_orig_index'] = df.index

        # Оставляем только строки с кодами
        df = df.dropna(subset=['Код работы'])
        df['Код работы'] = df['Код работы'].astype(str).str.strip()
        df = df[df['Код работы'] != ''].reset_index(drop=True)

        if df.empty:
            self._log("⚠️ Нет строк с кодами работ после очистки")
            return pd.DataFrame()

        self._log(f"   Строк после очистки: {len(df)}")

        # === [1] Отделяем строки без матча ===
        empty_match_mask = df['Смэтченная работа ВиКР'].fillna('').astype(str).str.strip() == ''
        no_match_df = df[empty_match_mask].copy()
        df_matched = df[~empty_match_mask].copy()

        self._log(f"   Без матча (проходят без агрегации): {len(no_match_df)}")
        self._log(f"   С матчем: {len(df_matched)}")

        # === [2] Нормализация ед. изм. перед сравнением ===
        df_matched['_vikr_unit'] = df_matched['Смэтченная работа ВиКР'].apply(self._extract_vikr_unit)

        vr_unit_norm = self._normalize_unit(df_matched['Ед. изм.'])
        vikr_unit_norm = self._normalize_unit(df_matched['_vikr_unit'])

        unit_mismatch_mask = (
                (vikr_unit_norm != '') &
                (vr_unit_norm != '') &
                (vr_unit_norm != vikr_unit_norm)
        )
        unit_mismatch_df = df_matched[unit_mismatch_mask].copy()
        df_to_agg = df_matched[~unit_mismatch_mask].copy()

        self._log(f"   Несовпадение ед. изм. (проходят без агрегации): {len(unit_mismatch_df)}")
        self._log(f"   Для агрегации: {len(df_to_agg)}")

        # 🔑 НАСТРОЙКА ВТОРОЙ КОЛОНКИ ("Наименование ВиКР")
        # 1. Без матча → "Не найдена подходящая"
        no_match_df['_vikr_name'] = "Не найдена подходящая"

        # 2. Ед. изм. не совпали → оригинальное название из ВР
        unit_mismatch_df['_vikr_name'] = unit_mismatch_df['Наименование ВР']

        # 3. Для агрегации → берём имя из справочника ВиКР (до первого "|")
        df_to_agg['_vikr_name'] = df_to_agg['Смэтченная работа ВиКР'].apply(
            lambda x: x.split('|')[0].strip() if isinstance(x, str) and '|' in x else str(x)
        )

        # Числовые колонки (защита от отсутствия)
        for col in ['Уверенность модели', 'Score эмбеддинга', 'Позиция в топ-5']:
            if col in df_to_agg.columns:
                df_to_agg[col] = pd.to_numeric(df_to_agg[col], errors='coerce').fillna(0)

        # === [3] Логика суммирования количеств ===
        def _agg_qty(series: pd.Series) -> float:
            """
            Умная агрегация:
            1. Все одинаковые → берём одно
            2. Макс == сумма остальных → берём макс (он уже включает остальные)
            3. Иначе → суммируем
            """
            vals = series.dropna().tolist()
            if not vals: return 0.0
            if len(vals) == 1: return vals[0]
            if len(set(vals)) == 1: return vals[0]

            max_val = max(vals)
            others_sum = sum(vals) - max_val
            if abs(others_sum - max_val) < 1e-6: return max_val

            return sum(vals)

        self._log("🧮 Агрегация данных...")
        self.progress_callback(1, 3, "📊 Агрегация")

        if not df_to_agg.empty:
            df_to_agg['group_key'] = (
                    df_to_agg['Раздел'].astype(str) + '||' +
                    df_to_agg['Подраздел'].astype(str) + '||' +
                    df_to_agg['Код работы'].astype(str)
            )

            agg_dict = {
                'Кол-во': _agg_qty,
                '№ п/п': lambda x: ', '.join(v for v in x.unique() if v),
                '_vikr_name': 'first',
                'Смэтченная работа ВиКР': 'first',
                'Код работы': 'first',
                'Раздел': 'first',
                'Подраздел': 'first',
                'Наименование ВР': lambda x: ' | '.join(x.unique()),
                '_orig_index': 'first'  # Для сортировки
            }

            if 'Ед. изм.' in df_to_agg.columns:
                agg_dict['Ед. изм.'] = lambda x: next((v for v in x if v.strip()), '')
            if 'Уверенность модели' in df_to_agg.columns:
                agg_dict['Уверенность модели'] = 'mean'
            if 'Score эмбеддинга' in df_to_agg.columns:
                agg_dict['Score эмбеддинга'] = 'mean'
            if 'Позиция в топ-5' in df_to_agg.columns:
                agg_dict['Позиция в топ-5'] = 'first'

            grouped = df_to_agg.groupby('group_key', sort=False).agg(agg_dict)
            grouped['Количество объединённых работ'] = df_to_agg.groupby('group_key', sort=False).size()
            aggregated_df = grouped.reset_index(drop=True)
        else:
            aggregated_df = pd.DataFrame()

        self.progress_callback(2, 3, "📊 Агрегация")

        # === Подготовка "сквозных" строк (без матча + несовпавшие ед. изм.) ===
        passthrough_parts = []
        for part_df in [no_match_df, unit_mismatch_df]:
            if not part_df.empty:
                pt = part_df.copy()
                pt['Количество объединённых работ'] = 1
                passthrough_parts.append(pt)

        self._log("📋 Формирование итоговой таблицы...")
        parts_to_concat = [aggregated_df] if not aggregated_df.empty else []
        parts_to_concat.extend(passthrough_parts)

        if not parts_to_concat:
            self._log("⚠️ Нет данных для формирования результата")
            return pd.DataFrame()

        final_df = pd.concat(parts_to_concat, ignore_index=True)

        # 🔑 ВОССТАНОВЛЕНИЕ ИСХОДНОГО ПОРЯДКА
        if '_orig_index' in final_df.columns:
            final_df = final_df.sort_values('_orig_index').reset_index(drop=True)
            final_df = final_df.drop(columns=['_orig_index'])

        # Удаляем служебные колонки
        for col in ['_vikr_unit', 'group_key']:
            if col in final_df.columns:
                final_df = final_df.drop(columns=[col])

        # Переименовываем для финального вывода
        final_df = final_df.rename(columns={
            '_vikr_name': 'Наименование ВиКР',
            'Наименование ВР': 'Наименования работ ВОР',
            '№ п/п': 'Объединённые № п/п'
        })

        # Порядок колонок
        col_order = [
            'Объединённые № п/п',
            'Наименование ВиКР',
            'Ед. изм.',
            'Кол-во',
            'Раздел',
            'Подраздел',
            'Код работы',
            'Смэтченная работа ВиКР',
            'Позиция в топ-5',
            'Уверенность модели',
            'Score эмбеддинга',
            'Наименования работ ВОР',
            'Количество объединённых работ',
        ]
        col_order = [c for c in col_order if c in final_df.columns]
        extra_cols = [c for c in final_df.columns if c not in col_order]
        final_df = final_df[col_order + extra_cols]

        self.progress_callback(3, 3, "📊 Агрегация")
        self._log(f"✅ Агрегация завершена: {len(final_df)} строк (из них агрегировано: {len(aggregated_df)})")
        return final_df

    @staticmethod
    def save_df_to_excel_with_highlights(df: pd.DataFrame, output_path: Union[str, Path]):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        output_path = Path(output_path)
        wb = Workbook()
        ws = wb.active

        headers = list(df.columns)
        ws.append(headers)

        qty_col_idx = headers.index('Кол-во') + 1 if 'Кол-во' in headers else None
        score_col_idx = headers.index('Score эмбеддинга') + 1 if 'Score эмбеддинга' in headers else None
        pos_col_idx = headers.index('Позиция в топ-5') + 1 if 'Позиция в топ-5' in headers else None

        warn_fill = PatternFill(start_color='FFFFCC00', end_color='FFFFCC00', fill_type='solid')

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            values = [None if pd.isna(row[col]) else row[col] for col in headers]
            ws.append(values)

            if qty_col_idx:
                cell = ws.cell(row=row_idx, column=qty_col_idx)
                try:
                    cell.value = float(row['Кол-во'])
                    # 0.########## : точка гарантирована, нули справа обрезаются, тип = число
                    cell.number_format = '0.##########'
                except (ValueError, TypeError):
                    pass

            needs_highlight = False
            if score_col_idx:
                try:
                    score_val = float(row.get('Score эмбеддинга', 0))
                    if 0 < score_val < 0.5:
                        needs_highlight = True
                except (ValueError, TypeError):
                    pass
            if pos_col_idx and not needs_highlight:
                try:
                    pos_val = float(row.get('Позиция в топ-5', 1))
                    if pos_val != 1 and pos_val > 0:
                        needs_highlight = True
                except (ValueError, TypeError):
                    pass

            if needs_highlight:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warn_fill

        for col_idx, header in enumerate(headers, 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(max(len(header) + 2, 10), 50)

        wb.save(str(output_path))

    def save_to_excel(self, df: pd.DataFrame, output_path: Union[str, Path]):
        self.save_df_to_excel_with_highlights(df, output_path)
        self._log(f"💾 Результат сохранён: {Path(output_path).name}")

    def _log(self, msg: str):
        if self._log_callback:
            self._log_callback(msg)
        else:
            print(msg)